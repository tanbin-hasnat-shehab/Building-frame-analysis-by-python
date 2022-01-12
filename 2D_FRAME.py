import time
from PIL import ImageTk,Image
import math
import os
import numpy as np
import matplotlib.pyplot as plt
import sympy as sy
from sympy import*
from tkinter import *
from openpyxl.workbook import *
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import colorchooser
from tkinter import ttk


def analyze(pxx,pyy,s_or_b):


    wb=load_workbook("data.xlsx")
    sheet=wb['Sheet']
    inertia_sheet=wb['inertia']
    omega_sheet=wb['omega']
    sup_x_sheet=wb['s_x']
    sup_y_sheet=wb['s_y']
    hz_left_sheet=wb['hz_loads_leftward']
    hz_right_sheet=wb['hz_loads_rightward']

    class Members:
        def __init__(self,x1,y1,x2,y2,length,w,I):
            self.x1=x1
            self.x2=x2
            self.y1=y1
            self.y2=y2
            self.length=length
            self.w=w
            self.I=I
            



    member_objects=[]
    for i in range(1,len(sheet['A'])+1):
        trial_obj=Members(sheet.cell(row=i,column=1).value , 
                    sheet.cell(row=i,column=2).value,
                    sheet.cell(row=i,column=3).value,
                    sheet.cell(row=i,column=4).value,
                    sheet.cell(row=i,column=5).value,
                    omega_sheet.cell(row=i,column=1).value,
                    inertia_sheet.cell(row=i,column=1).value)
        member_objects.append(trial_obj)




    for i in range(1,len(sheet['A'])+1):

        var1=0
        for j in range(6,100):
            if sheet.cell(row=i,column=j).value!=None:
                var1=var1+1
        pp=var1/3
        member_objects[i-1].no_p_loads=pp   
    #-----------------------
    for i in range(1,len(sheet['A'])+1):

        var1=0
        for j in range(1,100):
            if hz_right_sheet.cell(row=i,column=j).value!=None:
                var1=var1+1
        pp=var1/3
        member_objects[i-1].no_hz_p_loads=pp

    #------------------------------

    for i in range(0,len(member_objects)):
        if member_objects[i].y1==member_objects[i].y2:
            if member_objects[i].no_p_loads!=0:
                sum1=0
                sum2=0
                sum3=0
                sum4=0
                for j in range(0,int(member_objects[i].no_p_loads)):
                    pp=sheet.cell(row=i+1,column=6+j*3).value
                    bb=(member_objects[i].x2- sheet.cell(row=i+1,column=6+j*3+1).value  )
                    aa=sheet.cell(row=i+1,column=6+j*3+1).value - member_objects[i].x1 
                    ll=member_objects[i].length
                    sum1=sum1+pp*(bb**2)*aa/(ll**2)
                    sum2=sum2+pp*(bb**1)*(aa)**2/(ll**2)
                    sum3=sum3+pp*(bb**2)*(3*aa+bb)/(ll**3)
                    sum4=sum4+pp*(aa**2)*(aa+3*bb)/(ll**3)
                    
                member_objects[i].fem1=sum1+member_objects[i].w*(member_objects[i].length)**2/12
                member_objects[i].fem2=(sum2+member_objects[i].w*(member_objects[i].length)**2/12)*(-1)
                member_objects[i].fef1=sum3+member_objects[i].w*(member_objects[i].length)/2
                member_objects[i].fef2=sum4+member_objects[i].w*(member_objects[i].length)/2

            else:
                member_objects[i].fem1=member_objects[i].w*(member_objects[i].length)**2/12
                member_objects[i].fem2=(member_objects[i].w*(member_objects[i].length)**2/12)*(-1)
                member_objects[i].fef1=member_objects[i].w*(member_objects[i].length)/2
                member_objects[i].fef2=member_objects[i].w*(member_objects[i].length)/2



            
        else:
            
            if member_objects[i].no_hz_p_loads!=0:
                sum1=0
                sum2=0
                sum3=0
                sum4=0
                for j in range(0,int(member_objects[i].no_hz_p_loads)):
                    pp=hz_right_sheet.cell(row=i+1,column=1+j*3).value                  ###########################
                    bb=(member_objects[i].y2- hz_right_sheet.cell(row=i+1,column=1+j*3+2).value  )
                    aa=hz_right_sheet.cell(row=i+1,column=1+j*3+2).value - member_objects[i].y1 
                    ll=member_objects[i].length
                    sum1=sum1+pp*(bb**2)*aa/(ll**2)
                    sum2=sum2+pp*(bb**1)*(aa)**2/(ll**2)
                    sum3=sum3+pp*(bb**2)*(3*aa+bb)/(ll**3)
                    sum4=sum4+pp*(aa**2)*(aa+3*bb)/(ll**3)
                    
                member_objects[i].fem1=-sum1
                member_objects[i].fem2=sum2
                member_objects[i].fef1=-sum3
                member_objects[i].fef2=-sum4

            else:
                member_objects[i].fem1=0
                member_objects[i].fem2=0
                member_objects[i].fef1=0
                member_objects[i].fef2=0
    ##-----------------------------------------------------------------

    for i in range(0,len(member_objects)):
        print(f'member number {i}')
        print(f'forces ---- 1st one is {member_objects[i].fef1}  and 2nd is {member_objects[i].fef2}')
        print(f'moments-------1st one is {member_objects[i].fem1}  and 2nd is {member_objects[i].fem2}')
        print("\n")

        

    class Doki_points:
        def __init__(self,doki_x,doki_y):
            self.doki_x=doki_x
            self.doki_y=doki_y


    class All_points:
        def __init__(self,x,y):
            self.x=x
            self.y=y

    supports=[]
    for i in range(1,len(sup_x_sheet['A'])+1):
        s_trial=[sup_x_sheet.cell(row=i,column=1).value,sup_y_sheet.cell(row=i,column=1).value]
        supports.append(s_trial)

    trial_d_points=[]
    for i in range(0,len(member_objects)):
        if [member_objects[i].x1,member_objects[i].y1]   not in trial_d_points :
            t1=[member_objects[i].x1,member_objects[i].y1]
            trial_d_points.append(t1)
    for i in range(0,len(member_objects)):
        if [member_objects[i].x2,member_objects[i].y2]   not in trial_d_points :
            t2=[member_objects[i].x2,member_objects[i].y2]
            trial_d_points.append(t2)




    no_of_total_unique_points=len(trial_d_points)-len(supports)
    #print(no_of_total_unique_points)

    arr1=[]
    for i in range(0,len(trial_d_points)):
        arr1.append(trial_d_points[i][0])
    arr2=[]
    for i in range(0,len(trial_d_points)):
        arr2.append(trial_d_points[i][1])

    upper_left_point=[[min(arr1),min(arr2)]]
    lower_right_point=[[max(arr1),max(arr2)]]
    for i in supports:
        print(i)
    print(lower_right_point)

    points_arranged=[]
    for i in range(upper_left_point[0][1],lower_right_point[0][1]+1):
        for j in range(upper_left_point[0][0],lower_right_point[0][0]+1):
            if [j,i] in trial_d_points and [j,i] not in supports:
                points_arranged.append([j,i])

    points_arranged_all=[]
    for i in range(upper_left_point[0][1],lower_right_point[0][1]+1):
        for j in range(upper_left_point[0][0],lower_right_point[0][0]+1):
            if [j,i] in trial_d_points:
                points_arranged_all.append([j,i])




    print("----------------------------------------------")

    for i in range(0,len(member_objects)):
        member_objects[i].left_id=points_arranged.index([member_objects[i].x1,member_objects[i].y1])
        try:
            member_objects[i].right_id=points_arranged.index([member_objects[i].x2,member_objects[i].y2])
        except:
            member_objects[i].right_id=-1
    print("----------------------------------------------")
    for i in range(0,len(member_objects)):
        print(f'member_objects {i} left id={member_objects[i].left_id} and right id={member_objects[i].right_id}')
    print("----------------------------------------------")
    U = np.empty((len(points_arranged),len(points_arranged)),dtype=object)
    for i in range(0,len(points_arranged)):
        for j in range(0,len(points_arranged)):
            if i==j:
                a_val=[]
                for k in range(0,len(member_objects)):
                    numb=0
                    if j==member_objects[k].left_id or j==member_objects[k].right_id:

                        a_val.append(k)
                sum_4EibyL=0
                for m in range(0,len(a_val)):
                    sum_4EibyL=sum_4EibyL+4*member_objects[a_val[m]].I/member_objects[a_val[m]].length
                U[i,j]=sum_4EibyL
            else:
                desired=5000 
                for n in range(0,len(member_objects)):

                    if (member_objects[n].left_id==i and member_objects[n].right_id==j) or (member_objects[n].right_id==i and member_objects[n].left_id==j):
                        desired=n

                try:
                    vall=2*member_objects[desired].I/member_objects[desired].length
                    U[i,j]=vall
                    
                        
                except:
                    U[i,j]=0.0



    print("--------------------------------")

     

    tr111=[]
    for i in range(0,len(member_objects)):
        if member_objects[i].y1  not in tr111:
            tr111.append(member_objects[i].y1)
        if member_objects[i].y2 not in tr111:
            tr111.append(member_objects[i].y2)
    tr111.sort()
    f_doki_y=[]
    for i in range(0,len(tr111)-1):
        tr1t=tr111[i]
        f_doki_y.append(tr1t)
        
    no_of_force_doki=len(tr111)-1
    print("----------------------------------------------")

    class f_doki_upper_lower:
        def __init__(self,point_co_ord):
            self.point_co_ord=point_co_ord
            

    points_arranged_with_u_l=[]
    for i in range(0,len(points_arranged)):
        a5=f_doki_upper_lower(points_arranged[i])
        points_arranged_with_u_l.append(a5)


    for i in range(0,len(points_arranged_with_u_l)):
        for j in range(0,len(member_objects)):
            if member_objects[j].x1==member_objects[j].x2:
                if [member_objects[j].x1,member_objects[j].y1]==points_arranged_with_u_l[i].point_co_ord or [member_objects[j].x2,member_objects[j].y2]==points_arranged_with_u_l[i].point_co_ord:
                    if points_arranged_with_u_l[i].point_co_ord[1]<member_objects[j].y2:
                        points_arranged_with_u_l[i].lower_member_id=j
                    

                    if points_arranged_with_u_l[i].point_co_ord[1]==member_objects[j].y2:
                        points_arranged_with_u_l[i].upper_member_id=j
                    


    for i in range(0,len(points_arranged_with_u_l)):
        try:
            if points_arranged_with_u_l[i].upper_member_id>-5:
                pass
        except AttributeError:
            points_arranged_with_u_l[i].upper_member_id=-99

        try:
            if points_arranged_with_u_l[i].lower_member_id>-5:
                pass
        except AttributeError:
            points_arranged_with_u_l[i].lower_member_id=-99





    for i in range(0,len(points_arranged_with_u_l)):
        
        try:
            print(f'{i} th points upper member is {points_arranged_with_u_l[i].upper_member_id}')
        except:
            pass
        try:
            print(f'{i} th points lower member is {points_arranged_with_u_l[i].lower_member_id}')
        except:
            pass
        print("\n\n")


    #creating constant mat up to 8x8
    constant_mat=np.zeros(shape=(len(points_arranged),1))

    for i in range(0,len(points_arranged)):
        sss=[]
        sum_f_1=0
        for j in range(0,len(member_objects)):
            if i==member_objects[j].right_id and (points_arranged[i][1]==member_objects[j].y1) and (points_arranged[i][1]==member_objects[j].y2):
                sum_f_1=member_objects[j].fem2
        sum_f_2=0
        for j in range(0,len(member_objects)):
            if i==member_objects[j].left_id and (points_arranged[i][1]==member_objects[j].y1) and (points_arranged[i][1]==member_objects[j].y2):
                sum_f_2=member_objects[j].fem1
        
        sum_f_3=0
        for j in range(0,len(member_objects)):
            if i==member_objects[j].right_id and (points_arranged[i][0]==member_objects[j].x1) and (points_arranged[i][0]==member_objects[j].x2):
                sum_f_3=member_objects[j].fem2
        sum_f_4=0
        for j in range(0,len(member_objects)):
            if i==member_objects[j].left_id and (points_arranged[i][0]==member_objects[j].x1) and (points_arranged[i][0]==member_objects[j].x2):
                sum_f_4=member_objects[j].fem1
        #constant_mat[0,i]=sum_f_1+sum_f_2
        print(f'for point {i} s1={sum_f_1}  s2={sum_f_2}  s3={sum_f_3}  s4={sum_f_4}')
        constant_mat[i]=sum_f_1+sum_f_2+sum_f_3+sum_f_4

    #3c 97 10 ba 15 c9

    print("----------------------------------------------")




    print("---------------------------------@@@@@@@@@@@@@@")
    def f_doki_stiffness_bottom(i,j):
        where_applied=j
        where_happened=i
        
        tr_val=0
        
        print(f'where happend {where_happened} and where where_applied is {where_applied}')
        if f_doki_y[where_happened]==points_arranged_with_u_l[where_applied].point_co_ord[1]:

            if points_arranged_with_u_l[where_applied].upper_member_id==-99 and points_arranged_with_u_l[where_applied].lower_member_id!=-99:
                tr_val=6*member_objects[points_arranged_with_u_l[where_applied].lower_member_id].I/member_objects[points_arranged_with_u_l[where_applied].lower_member_id].length**2
                print(f'this is 1st and tr_val is {tr_val}')
            
            if points_arranged_with_u_l[where_applied].upper_member_id!=-99 and points_arranged_with_u_l[where_applied].lower_member_id!=-99: 
                tr_val=6*member_objects[points_arranged_with_u_l[where_applied].lower_member_id].I/member_objects[points_arranged_with_u_l[where_applied].lower_member_id].length**2-6*member_objects[points_arranged_with_u_l[where_applied].upper_member_id].I/member_objects[points_arranged_with_u_l[where_applied].upper_member_id].length**2
                
                print(f'this is 2nd and tr_val is {tr_val}')

        try:
            if f_doki_y[where_happened]==member_objects[points_arranged_with_u_l[where_applied].upper_member_id].y1:
                tr_val=6*member_objects[points_arranged_with_u_l[where_applied].upper_member_id].I/member_objects[points_arranged_with_u_l[where_applied].upper_member_id].length**2
        except:
            pass
        try:
            if f_doki_y[where_happened]==member_objects[points_arranged_with_u_l[where_applied].lower_member_id].y2:
                tr_val=-6*member_objects[points_arranged_with_u_l[where_applied].lower_member_id].I/member_objects[points_arranged_with_u_l[where_applied].lower_member_id].length**2
        except:
            pass



        return tr_val
    print("----------------------------------------------------")

    bottom_mat = np.empty([no_of_force_doki,len(points_arranged)])

    for i in range(0,no_of_force_doki):
        for j in range(0,len(points_arranged)):
            bottom_mat[i,j] = f_doki_stiffness_bottom(i,j)
            



    print("---------------------------------@@@@@@@@@@@@@@")
    def f_doki_stiffness_right(i,j):
        where_applied=i
        where_happened=j
        
        tr_val=0
        
        #print(f'where happend {where_happened} and where where_applied is {where_applied}')
        if f_doki_y[where_happened]==points_arranged_with_u_l[where_applied].point_co_ord[1]:

            if points_arranged_with_u_l[where_applied].upper_member_id==-99 and points_arranged_with_u_l[where_applied].lower_member_id!=-99:
                tr_val=6*member_objects[points_arranged_with_u_l[where_applied].lower_member_id].I/member_objects[points_arranged_with_u_l[where_applied].lower_member_id].length**2
                print(f'this is 1st and tr_val is {tr_val}')
            
            if points_arranged_with_u_l[where_applied].upper_member_id!=-99 and points_arranged_with_u_l[where_applied].lower_member_id!=-99: 
                tr_val=6*member_objects[points_arranged_with_u_l[where_applied].lower_member_id].I/member_objects[points_arranged_with_u_l[where_applied].lower_member_id].length**2-6*member_objects[points_arranged_with_u_l[where_applied].upper_member_id].I/member_objects[points_arranged_with_u_l[where_applied].upper_member_id].length**2
                
                print(f'this is 2nd and tr_val is {tr_val}')

        try:
            if f_doki_y[where_happened]==member_objects[points_arranged_with_u_l[where_applied].upper_member_id].y1:
                tr_val=6*member_objects[points_arranged_with_u_l[where_applied].upper_member_id].I/member_objects[points_arranged_with_u_l[where_applied].upper_member_id].length**2
        except:
            pass
        try:
            if f_doki_y[where_happened]==member_objects[points_arranged_with_u_l[where_applied].lower_member_id].y2:
                tr_val=-6*member_objects[points_arranged_with_u_l[where_applied].lower_member_id].I/member_objects[points_arranged_with_u_l[where_applied].lower_member_id].length**2
        except:
            pass



        return tr_val
    print("----------------------------------------------------")

    right_mat = np.empty([len(points_arranged),no_of_force_doki])

    for i in range(0,no_of_force_doki):
        for j in range(0,len(points_arranged)):
            right_mat[j,i] = f_doki_stiffness_right(j,i)





    print("---------------------------------@@@@@@@@@@@@@@")
    def f_doki_stiffness_lower_right(i,j):
        where_applied=j
        where_happened=i
        
        
        sum_ff=0
        if where_happened==where_applied:
            for k in range(0,len(points_arranged_with_u_l)):
                if f_doki_y[where_happened]==points_arranged_with_u_l[k].point_co_ord[1]:
                    if points_arranged_with_u_l[k].upper_member_id!=-99:
                        sum_ff=sum_ff+12*member_objects[points_arranged_with_u_l[k].upper_member_id].I/member_objects[points_arranged_with_u_l[k].upper_member_id].length**3+12*member_objects[points_arranged_with_u_l[k].lower_member_id].I/member_objects[points_arranged_with_u_l[k].lower_member_id].length**3
                    if points_arranged_with_u_l[k].upper_member_id==-99:
                        sum_ff=sum_ff+12*member_objects[points_arranged_with_u_l[k].lower_member_id].I/member_objects[points_arranged_with_u_l[k].lower_member_id].length**3
        
        if where_happened<where_applied and (where_applied-where_happened)==1 and where_happened!=len(f_doki_y)-1:
            for k in range(0,len(points_arranged_with_u_l)):
                if f_doki_y[where_happened]<points_arranged_with_u_l[k].point_co_ord[1] and f_doki_y[where_happened+1]==points_arranged_with_u_l[k].point_co_ord[1]:
                    try:
                        sum_ff=sum_ff-12*member_objects[points_arranged_with_u_l[k].upper_member_id].I/member_objects[points_arranged_with_u_l[k].upper_member_id].length**3
                    except:
                        pass

        if where_happened>where_applied and (where_happened-where_applied)==1:
            
            for k in range(0,len(points_arranged_with_u_l)):
                if f_doki_y[where_applied+1]==points_arranged_with_u_l[k].point_co_ord[1]:
            
                    try:
                        sum_ff=sum_ff-12*member_objects[points_arranged_with_u_l[k].upper_member_id].I/member_objects[points_arranged_with_u_l[k].upper_member_id].length**3
                    except:
                        pass

        return sum_ff


        
    print("----------------------------------------------------")

    lower_right_mat = np.empty([no_of_force_doki,no_of_force_doki])

    for i in range(0,no_of_force_doki):
        for j in range(0,no_of_force_doki):
            lower_right_mat[i,j] = f_doki_stiffness_lower_right(i,j)



    print("----------------------------------------------------")
    print("----------------------------------------------------")
    print("------------------stiffness mat part 1----------------------------")     
    for i in range(0,len(U)):
        for j in range(0,len(U)):
            print(f'[{i},{j}] = {U[i,j]}')
    print("----------------------------------------------------")
    print("----------------------------------------------------")
    print("-------------constant mat--------------------------------")  
    print(constant_mat)
    print("----------------------------------------------------")
    print("----------------------------------------------------")
    print("-------------bottom_mat--------------------------------")    
    print(bottom_mat)
    print("----------------------------------------------------")
    print("----------------------------------------------------")
    print("---------------right mat-----------------------------")          
    print(right_mat)
    print("----------------------------------------------------")
    print("----------------------------------------------------")
    print("-----------------lower_right_mat--------------------------") 
            
    print(lower_right_mat)



    lower_final_mat=np.concatenate((bottom_mat,lower_right_mat),axis=1)
    upper_final_mat=np.concatenate((U,right_mat),axis=1)
    mother=np.concatenate((upper_final_mat,lower_final_mat))
    print("----------------------------------------------------")
    print("----------------------------------------------------")
    print("----------------------------------------------------")
    print("----------------------------------------------------")
    print("----------------------------------------------------")
    print("----------------------mother------------------------------")
    print(mother)

    print("----------------------------------------------------")
    print("-------------------- constants mat-------------------------------")
    constant_mat_2=np.zeros(shape=(len(f_doki_y),1))
    for i in range(0,len(f_doki_y)):
        s_1=0
        for j in range(0,len(member_objects)):
            if member_objects[j].x1==member_objects[j].x2:
                if f_doki_y[i]==member_objects[j].y1:
                    s_1=s_1+member_objects[j].fef1
                if f_doki_y[i]==member_objects[j].y2:
                    s_1=s_1+member_objects[j].fef2

        constant_mat_2[i]=s_1
    mother_right=np.concatenate((constant_mat,constant_mat_2))
    mother=sy.Matrix(mother)
    mother_right=sy.Matrix(mother_right)
    print(mother_right)
    print("---------soln--------")

    sloution=mother.inv()*(-mother_right)
    print(sloution)
    deflection_pointer=open('Deflections at joints.txt','a')
    deflection_pointer.write("--------------------------- "+"\n")
    for i in range(0,len(sloution)):
        if i<len(U):
            deflection_pointer.write("rotation "+str(sloution[i])+"\n")
        else:
            deflection_pointer.write("displacement "+str(sloution[i])+"\n")

    deflection_pointer.close()


    for i in range(0,len(member_objects)):

        if member_objects[i].y1==member_objects[i].y2:

            member_objects[i].fem1=member_objects[i].fem1+4/member_objects[i].length*sloution[member_objects[i].left_id]+2/member_objects[i].length*sloution[member_objects[i].right_id]
            member_objects[i].fem2=member_objects[i].fem2+4/member_objects[i].length*sloution[member_objects[i].right_id]+2/member_objects[i].length*sloution[member_objects[i].left_id]
        
    ##--------------vertical---------------
        if member_objects[i].x1==member_objects[i].x2:
            trial_contribution_fem1=0
            
            for j in range(0,len(f_doki_y)):
                if member_objects[i].y1==f_doki_y[j]:
                    upper_doki=j
                    trial_contribution_fem1=trial_contribution_fem1+6/member_objects[i].length**2*sloution[len(U)+upper_doki]
                    
                
                if member_objects[i].y2==f_doki_y[j]:
                    lower_doki=j
                    
                    
            
            
            if member_objects[i].y2 not in f_doki_y:
                pass
            else:
                trial_contribution_fem1=trial_contribution_fem1-6/member_objects[i].length**2*sloution[len(U)+lower_doki]
            

            for j in range(0,len(points_arranged_with_u_l)):
                if [member_objects[i].x1,member_objects[i].y1]==points_arranged_with_u_l[j].point_co_ord:
                    upper_doki=j
                    trial_contribution_fem1=trial_contribution_fem1+4/member_objects[i].length*sloution[upper_doki]
                if [member_objects[i].x2,member_objects[i].y2]==points_arranged_with_u_l[j].point_co_ord:
                    lower_doki=j
                    trial_contribution_fem1=trial_contribution_fem1+2/member_objects[i].length*sloution[lower_doki]


            member_objects[i].fem1=member_objects[i].fem1+ trial_contribution_fem1

    #_------------------------------------
        if member_objects[i].x1==member_objects[i].x2:
            trial_contribution_fem2=0
            
            for j in range(0,len(f_doki_y)):
                if member_objects[i].y1==f_doki_y[j]:
                    upper_doki=j
                    trial_contribution_fem2=trial_contribution_fem2+6/member_objects[i].length**2*sloution[len(U)+upper_doki]
                    
                
                if member_objects[i].y2==f_doki_y[j]:
                    lower_doki=j
                    trial_contribution_fem2=trial_contribution_fem2-6/member_objects[i].length**2*sloution[len(U)+lower_doki]

            for j in range(0,len(points_arranged_with_u_l)):
                if [member_objects[i].x1,member_objects[i].y1]==points_arranged_with_u_l[j].point_co_ord:
                    upper_doki=j
                    trial_contribution_fem2=trial_contribution_fem2+2/member_objects[i].length*sloution[upper_doki]
                if [member_objects[i].x2,member_objects[i].y2]==points_arranged_with_u_l[j].point_co_ord:
                    lower_doki=j
                    trial_contribution_fem2=trial_contribution_fem2+4/member_objects[i].length*sloution[lower_doki]


            member_objects[i].fem2=member_objects[i].fem2+ trial_contribution_fem2

    #_------------------------------------
    for i in range(0,len(member_objects)):
        print(f'member number {i} : left moment is {member_objects[i].fem1} and right moment is {member_objects[i].fem2}')

    #_------------------------------------

    print("=---------------------------------------")
    print(member_objects[1].no_p_loads)

    def horizontal_membersolve(a,s_or_b):
        Ma=member_objects[a].fem1
        Mb=member_objects[a].fem2
        l=member_objects[a].length
        omega=member_objects[a].w

        Ra_1st=(Ma+Mb+omega*l**2/2)/l
        additional=0
        sum_of_p_loads=0
        if member_objects[a].no_p_loads==0:
            pass    
        else:
            for i in range(0,int(member_objects[a].no_p_loads)):
                additional=additional+sheet.cell(row=a+1,column=6+i*3).value*(member_objects[a].x2-sheet.cell(row=a+1,column=7+i*3).value)
                sum_of_p_loads=sum_of_p_loads+sheet.cell(row=a+1,column=6+i*3).value
        Ra=Ra_1st+additional/l
        Rb=(omega*l+sum_of_p_loads)-Ra
        moment_points=[]
        moment_points_x=[]
        shear_force_points=[]
        shear_force_points_x=[]
        for i in range(0,l+1):
            if i==0:
                moment_points.append(0)
                moment_points_x.append(0)

                moment_points.append(-Ma)
                
                moment_points_x.append(0)


                
            else:
                p_load_contribution=0
                if member_objects[a].no_p_loads!=0:
                    for j in range(0,int(member_objects[a].no_p_loads)):
                        if sheet.cell(row=a+1,column=7+j*3).value<i+member_objects[a].x1:
                            p_load_contribution=p_load_contribution+sheet.cell(row=a+1,column=6+j*3).value*(i+member_objects[a].x1-sheet.cell(row=a+1,column=7+j*3).value)

                moment_points.append(-Ma-omega*i**2/2-p_load_contribution+Ra*i)
                moment_points_x.append(i)
        
        moment_points_x.append(l)
        moment_points.append(0)
        moment_points_x.append(0)
        moment_points.append(0)


        #for i in moment_points:
            #print(i)
        print("----------------------------------")

        for i in range(0,l+1):
            if i==0:
                
                shear_force_points.append(0)
                shear_force_points_x.append(0)
                shear_force_points.append(Ra)
                shear_force_points_x.append(0)
                
                



            else:
                point_load_contribution_on_sfd=0
                new_load=0
                if member_objects[a].no_p_loads!=0:
                    
                    for j in range(0,int(member_objects[a].no_p_loads)):
                        if i+member_objects[a].x1>sheet.cell(row=a+1,column=7+j*3).value:
                            point_load_contribution_on_sfd=point_load_contribution_on_sfd+sheet.cell(row=a+1,column=6+j*3).value
                        if i+member_objects[a].x1==sheet.cell(row=a+1,column=7+j*3).value:
                            new_load=sheet.cell(row=a+1,column=6+j*3).value
                            
                        

                shear_force_points.append(Ra-omega*i-point_load_contribution_on_sfd)
                shear_force_points_x.append(i)
                shear_force_points.append(Ra-omega*i-point_load_contribution_on_sfd-new_load)
                shear_force_points_x.append(i)
                
        shear_force_points.append(0)
        shear_force_points_x.append(l)
        shear_force_points.append(0)
        shear_force_points_x.append(0)          

        print("Moments points")
        for i in range(0,len(moment_points)):
            print(f'{moment_points_x[i]}    {moment_points[i]}')
        print("shear_force_points")
        for i in range(0,len(shear_force_points)):
            print(f'{shear_force_points_x[i]}   {shear_force_points[i]}')
        
        for i in range(0,len(moment_points)):
            moment_points[i]=-moment_points[i]


        if s_or_b=='s':
            plt.plot(shear_force_points_x,shear_force_points)
            plt.xlabel('SHEAR FORCE IN KIP')
            
            plt.show()
            
        if s_or_b=='b':
            plt.plot(moment_points_x,moment_points)
            plt.xlabel('BENDING MOMENT IN KIP-ft')
            plt.show()









    def vertical_membersolve(a,s_or_b):
        Ma=member_objects[a].fem1
        Mb=member_objects[a].fem2
        l=member_objects[a].length
        

        Ra_1st=(Ma+Mb)/l
        additional=0
        sum_of_p_loads=0
        if member_objects[a].no_hz_p_loads==0:
            pass    
        else:
            for i in range(0,int(member_objects[a].no_hz_p_loads)):
                additional=additional+hz_right_sheet.cell(row=a+1,column=1+i*3).value*(hz_right_sheet.cell(row=a+1,column=3+i*3).value-member_objects[a].y1)
                sum_of_p_loads=sum_of_p_loads+hz_right_sheet.cell(row=a+1,column=1+i*3).value
        Ra=Ra_1st+additional/l
        Rb=(sum_of_p_loads)-Ra
        cc=Rb
        Rb=Ra
        Ra=cc

        



        moment_points=[]
        moment_points_x=[]
        shear_force_points=[]
        shear_force_points_x=[]
        for i in range(0,l+1):
            if i==0:
                moment_points_x.append(l)
                moment_points.append(0)

                moment_points_x.append(l)
                moment_points.append(Ma)




                
            else:
                p_load_contribution=0
                if member_objects[a].no_hz_p_loads!=0:
                    for j in range(0,int(member_objects[a].no_hz_p_loads)):
                        if hz_right_sheet.cell(row=a+1,column=3+j*3).value<(i+member_objects[a].y1):
                            p_load_contribution=p_load_contribution+hz_right_sheet.cell(row=a+1,column=1+j*3).value*((i+member_objects[a].y1)-hz_right_sheet.cell(row=a+1,column=3+j*3).value)
                #print(f'i = {i}  then p_load_contribution is {p_load_contribution}')
                moment_points.append(Ma-p_load_contribution+Ra*i)
                moment_points_x.append(l-i)
        moment_points_x.append(0)
        moment_points.append(0)
        moment_points_x.append(l)
        moment_points.append(0)

        print("----------------------------------")
        #for i in moment_points:
            #print(i)
        print("----------------------------------")

        for i in range(0,l+1):
            if i==0:
                

                shear_force_points_x.append(l)
                shear_force_points.append(0)

                shear_force_points_x.append(l)
                shear_force_points.append(Ra)


            else:
                point_load_contribution_on_sfd=0
                new_load=0
                if member_objects[a].no_hz_p_loads!=0:
                    
                    for j in range(0,int(member_objects[a].no_hz_p_loads)):
                        if i+member_objects[a].y1>hz_right_sheet.cell(row=a+1,column=3+j*3).value:
                            point_load_contribution_on_sfd=point_load_contribution_on_sfd+hz_right_sheet.cell(row=a+1,column=1+j*3).value
                        if i+member_objects[a].y1==hz_right_sheet.cell(row=a+1,column=3+j*3).value:
                            new_load=hz_right_sheet.cell(row=a+1,column=1+j*3).value
                            
                        

                shear_force_points.append(Ra-point_load_contribution_on_sfd)
                shear_force_points_x.append(l-i)

                shear_force_points.append(Ra-point_load_contribution_on_sfd-new_load)
                shear_force_points_x.append(l-i)
        shear_force_points_x.append(0)
        shear_force_points.append(0)
        shear_force_points_x.append(l)
        shear_force_points.append(0)
        print("Moments points")
        for i in range(0,len(moment_points)):
            print(f'{moment_points_x[i]}    {moment_points[i]}')
        print("shear_force_points")
        for i in range(0,len(shear_force_points)):
            print(f'{shear_force_points_x[i]}   {shear_force_points[i]}')
        
        for i in range(0,len(shear_force_points)):
            shear_force_points[i]=-shear_force_points[i]
        if s_or_b=='s':
            plt.plot(shear_force_points,shear_force_points_x)
            plt.xlabel('SHEAR FORCE IN KIP')
            plt.show()
            
        if s_or_b=='b':
            plt.plot(moment_points,moment_points_x)
            plt.xlabel('BMD IN KIP-ft')
            plt.show()

                
                
                


        

    def draw(member_asked,s_or_b):
        if member_objects[member_asked].x1==member_objects[member_asked].x2:
            vertical_membersolve(member_asked,s_or_b)
        if member_objects[member_asked].y1==member_objects[member_asked].y2:
            horizontal_membersolve(member_asked,s_or_b)


    pxx=pxx/10
    pyy=pyy/10
    print(f'pxx={pxx} and pyy= {pyy}')
    min_dist=[]
    for i in range(0,len(member_objects)):
        min_dist.append(  round(math.sqrt((pxx-member_objects[i].x1)**2+(pyy-member_objects[i].y1)**2) + math.sqrt((pxx-member_objects[i].x2)**2+(pyy-member_objects[i].y2)**2)) )
    miin=min(min_dist)
    member_asked=min_dist.index(miin)
    print("member_asked = ",member_asked)
    draw(member_asked,s_or_b)










































####################---------------------------------------   UI    -----------------------------------------------#####################


























try:
    os.remove("data.xlsx")
except:

    pass
def end_line_command(event):
    my_canvas.unbind('<Button-1>')
    
#################   
def line_draw(event):

    my_canvas.unbind('<Motion>')
    def add_point_load(event):
        def send_point_load_value_to_excel(event):
            
            p_entry.config(state=DISABLED)
            p_entry.unbind('<Return>')
            filled_row=0
            for i in range(1,1000):
                if sheet.cell(row=counter,column=i).value!=None:
                    filled_row+=1   
            print(filled_row) 
            print(counter)
            print("x = ",aa)
            print("y = ",bb)
            
            sheet.cell(row=counter,column=filled_row+1).value=float(p_entry.get())
            sheet.cell(row=counter,column=filled_row+2).value=round(aa)
            sheet.cell(row=counter,column=filled_row+3).value=round(bb)
            wb.save("data.xlsx")
            
        global point_load_counter
        initial_p_load = StringVar( value='0')
        
        p_entry = Entry(my_canvas,textvariable=initial_p_load)
        my_canvas.create_window(event.x+o_x, event.y-70+o_y, window=p_entry, height=15, width=25)
        
        
        
        
        my_canvas.create_image((event.x+o_x,event.y-30+o_y) , image=point_load_image)
        
        
        aa=(event.x+o_x)/10
        bb=(event.y+o_y)/10
        p_entry.bind('<Return>',send_point_load_value_to_excel)
        point_load_counter+=1
        ##-------------------------------------------------------
        
        
    def add_point_load_h_right(event):
        def send_point_load_value_to_excel_h_right(event):
            p_entry.config(state=DISABLED)
            p_entry.unbind('<Return>')
            filled_row=0
            for i in range(1,1000):
                if hz_loads_rightward.cell(row=counter,column=i).value!=None:
                    filled_row+=1   
            print(filled_row) 
            print(counter)
            print("x = ",aa)
            print("y = ",bb)
            
            hz_loads_rightward.cell(row=counter,column=filled_row+1).value=float(p_entry.get())
            hz_loads_rightward.cell(row=counter,column=filled_row+2).value=round(aa)
            hz_loads_rightward.cell(row=counter,column=filled_row+3).value=round(bb)
            wb.save("data.xlsx")


        #------------------------------------------------------
        global point_load_counter
        initial_p_load = StringVar( value='0')
        
        p_entry = Entry(my_canvas,textvariable=initial_p_load)
        my_canvas.create_window(event.x-70+o_x, event.y+o_y, window=p_entry, height=15, width=25)
        
        
        
        
        my_canvas.create_image((event.x-30+o_x,event.y+o_y) , image=point_load_image_h_right)
        
        
        aa=(event.x+o_x)/10
        bb=(event.y+o_y)/10
        p_entry.bind('<Return>',send_point_load_value_to_excel_h_right)

        
    #----------------------------------------------    
    def add_point_load_h_left(event):

        def send_point_load_value_to_excel_h_left(event):
            filled_row=0
            for i in range(1,1000):
                if hz_loads_leftward.cell(row=counter,column=i).value!=None:
                    filled_row+=1   
            print(filled_row) 
            print(counter)
            print("x = ",aa)
            print("y = ",bb)
            
            hz_loads_leftward.cell(row=counter,column=filled_row+1).value=float(p_entry.get())
            hz_loads_leftward.cell(row=counter,column=filled_row+2).value=round(aa)
            hz_loads_leftward.cell(row=counter,column=filled_row+3).value=round(bb)
            wb.save("data.xlsx")

        ####------------------------------------------------
        global point_load_counter
        initial_p_load = StringVar( value='0')
        
        p_entry = Entry(my_canvas,textvariable=initial_p_load)
        my_canvas.create_window(event.x+70+o_x, event.y+o_y, window=p_entry, height=15, width=25)
        
        
        
        
        my_canvas.create_image((event.x+30+o_x,event.y+o_y) , image=point_load_image_h_left)
        
        
        aa=round((event.x+o_x)/off)*off/10
        bb=round((event.y+o_y)/off)*off/10
        p_entry.bind('<Return>',send_point_load_value_to_excel_h_left) 
    #--------------------------------------------------------- 
    global show_ord,show_window,color_p
    show_ord=Label(root,text='hi ',bg='#15ebe7' )
    
    #show_window=my_canvas.create_window(50,0,window=show_ord,anchor=NW)  
   
    def show_co_ordinate(event):
        

        global show_ord,show_window
        if sheet.cell(row=counter,column=1).value==sheet.cell(row=counter,column=3).value:
            show_ord.config(text=f'distance = {round(((event.y+o_y)/10-sheet.cell(row=counter,column=2).value))}')
            
            show_window=my_canvas.create_window(event.x+o_x-50,event.y+o_y-50,window=show_ord)
            
            
           
            
            
        
        if sheet.cell(row=counter,column=2).value==sheet.cell(row=counter,column=4).value: 
            show_ord.config(text=f'distance = {round(((event.x+o_x)/10-sheet.cell(row=counter,column=1).value))}')
           
            show_window=my_canvas.create_window(event.x+o_x,event.y+o_y-60,window=show_ord)
            
          
           
    def add_point_load_h_left_1(event):

        my_canvas.unbind('<Button-1>')
        my_canvas.unbind('<Button-3>')
        my_canvas.unbind('<Motion>')
        my_canvas.delete(show_window)
    def add_point_load_binding():
        global point_load_counter
        point_load_counter=1
        my_canvas.unbind('<Button-1>')
        my_canvas.bind('<Button-1>',add_point_load)
        my_canvas.bind('<Button-2>',add_point_load_h_left_1)
        my_canvas.bind('<Button-3>',add_point_load_h_right)
        my_canvas.bind('<Motion>',show_co_ordinate)
       
        
        
        
    
    def delete_line(event):
        global counter
        try:
            my_canvas.delete(my_line)
            counter-=1
        except:
            pass
    global c,counter,x1,y1,x2,y2
    if c==0:
        '''global x1
        global y1'''
        x1=event.x+o_x
        y1=event.y+o_y
        print(f'x1 ={x1} event.x={event.x} off={off} gr={gr}')
        r=3
        my_canvas.create_rectangle(round(x1/off)*off-r,round(y1/off)*off-r,round(x1/off)*off+r,round(y1/off)*off+r,fill="blue")
        c=1

    else:
        #global counter
        x2=event.x+o_x
        y2=event.y+o_y
        x1_def=x1
        x2_def=x2
        y1_def=y1
        y2_def=y2
        if round(y1/off)==round(y2/off):
            if round(x2/off)<round(x1/off):
                var1=x1
                x1=x2
                x2=var1
        if round(x1/off)==round(x2/off):
            if round(y2/off)<round(y1/off):
                var2=y1
                y1=y2
                y2=var2    
        
        x1=round(x1/off)*off
        x2=round(x2/off)*off
        y1=round(y1/off)*off
        y2=round(y2/off)*off
        print(f'o_x={o_x}  o_y={o_y}')
        my_line=my_canvas.create_line(x1,y1,x2,y2,width=3,fill='green')
        
        def send_dist_load_value_to_excel(event):
           
            
            
            omega_sheet.cell(row=counter,column=1).value=float(w_entry.get())
            inertia_sheet.cell(row=counter,column=1).value=float(inertia_entry.get())
            
        
        ####w
        initial_w_load = StringVar( value='0')
        w_entry = Entry(my_canvas,textvariable=initial_w_load)
        w_entry.bind('<Return>',send_dist_load_value_to_excel)
        #my_canvas.create_window((x1+x2)/2,(y1+y2)/2-20, window=w_entry, height=15, width=15)
        ####I
        initial_inertia = StringVar( value='1')
        inertia_entry = Entry(my_canvas,textvariable=initial_inertia)
        inertia_entry.bind('<Return>',send_dist_load_value_to_excel)
        if y1==y2:
            my_canvas.create_window((x1_def+x2_def)/2,(y1_def+y2_def)/2+20, window=inertia_entry, height=15, width=25)
    
            my_canvas.create_text((x1_def+x2_def)/2-20,(y1_def+y2_def)/2+20,text='I=')
        else:
            my_canvas.create_window((x1_def+x2_def)/2+10*3,(y1_def+y2_def)/2, window=inertia_entry, height=15, width=25)
    
            my_canvas.create_text((x1_def+x2_def)/2-20+10*3,(y1_def+y2_def)/2,text='I=')

        
        
        for i in range(0,int((math.sqrt((x2-x1)**2+(y2-y1)**2))/10)):
            if round(y1/off)==round(y2/off):
                my_canvas.create_line(x1+i*10,y1,x1+i*10,y1-15,width=1,fill='red')
                my_canvas.create_line(x1,y1-15,x2,y2-15,width=1,fill='red')
                my_canvas.create_window((x1+x2)/2,(y1+y2)/2-20, window=w_entry, height=15, width=25)
            
        
        counter+=1
        
        c=0
        
        x1=(round(x1/off))*off/10
        y1=(round(y1/off))*off/10
        x2=(round(x2/off))*off/10
        y2=(round(y2/off))*off/10
        length=math.sqrt((x1-x2)**2+(y1-y2)**2)
        
                
        sheet.cell(row=counter,column=1).value=x1
        sheet.cell(row=counter,column=2).value=y1
        sheet.cell(row=counter,column=3).value=x2
        sheet.cell(row=counter,column=4).value=y2
        sheet.cell(row=counter,column=5).value=length
        wb.save('data.xlsx')

        
        
        print("member no "+str(counter)+" = "+str(x1)+","+str(y1)+"  and  "+str(x2)+","+str(y2)+"   and length = "+str(length))
   
    point_load_image = PhotoImage(file='beams/p2.png')
    point_load_image_h_left = PhotoImage(file='beams/p2_h_left.png')
    point_load_image_h_right = PhotoImage(file='beams/p2_h_right.png')

    
    fixed_supp_image=PhotoImage(file='beams/fixed.png')
    hinged_supp_image=PhotoImage(file='beams/hinged.png')

    
    
    
    #dist_load_img = PhotoImage(file='beams/dis.png')
    my_canvas.bind('<Double-1>',end_line_command)
    my_canvas.bind('<Button-3>',delete_line)
    p_lo=Button(root,command=add_point_load_binding,image=point_btn_image,borderwidth=1,height=50,width=50)
    p_lo.place(x=0,y=110)
    
    command_label3=Label(root,text='Add point load downward or rightward',font=('Helvetica',20),bg='#15ebe7')
    p_lo.bind('<Enter>',lambda x:

                                    command_label3.place(x=50,y=110)
                                    
                                        )
    p_lo.bind('<Leave>',lambda x:
                                    command_label3.place_forget()
                                        )

    ##################


    #my_canvas.create_window(0,0,height=30,width=70,window=p_lo)
    def save_all():
        my_canvas.unbind('<Button-1>')
        my_canvas.unbind('<Button-3>')
        wb.save("data.xlsx")
    
    
    save_btn=Button(root,command=save_all,image=save_btn_image,borderwidth=1,height=50,width=80)
    save_btn.place(x=140,y=0)


    ###############
    
    command_label11=Label(root,text='Build',font=('Helvetica',20),bg='#15ebe7')
    save_btn.bind('<Enter>',lambda x:

                                    command_label11.place(x=140,y=100)
                                    
                                        )
    save_btn.bind('<Leave>',lambda x:
                                    command_label11.place_forget()
                                        )

    ##################

    #my_canvas.create_window(0,75,height=30,width=70,window=save_btn)

    def add_f_s(event):
        support_pos_x=round((event.x+o_x)/off)*off
        support_pos_y=round((event.y+o_y)/off)*off
        sup_x = [(support_pos_x)/10]
        sup_y = [(support_pos_y)/10]
        
        global support_type
        if support_type=='fixed':
            image_label1=Label(image=fixed_supp_image)
            sup_x_sheet.append(sup_x)
            sup_y_sheet.append(sup_y)

        if support_type=='hinged':
            image_label1=Label(image=hinged_supp_image)    
        my_canvas.create_window(support_pos_x,support_pos_y+12,window=image_label1,width=50,height=25)


    def add_fixed_support(text1):
        my_canvas.unbind('<Button-1>')
        global support_type
        if text1=='fixed':
            support_type="fixed"
            my_canvas.bind('<Button-1>', add_f_s)
        else:
            support_type="hinged"
            my_canvas.bind('<Button-1>',add_f_s)

    #fixed_btn_image=PhotoImage(file='beams/fix_3.png')
    #hinged_btn_image=PhotoImage(file='beams/hin_1.png')   
    fixed_support_btn=Button(root,image=fixed_btn_image,borderwidth=1,height=50,width=50,command=lambda : add_fixed_support("fixed"))
    fixed_support_btn.place(x=0,y=160)
    
    command_label4=Label(root,text='Add Fixed support',font=('Helvetica',20),bg='#15ebe7')
    fixed_support_btn.bind('<Enter>',lambda x:

                                    command_label4.place(x=50,y=160)
                                    
                                        )
    fixed_support_btn.bind('<Leave>',lambda x:
                                    command_label4.place_forget()
                                        )

    ##################



    hinge_support=Button(root,text='',image=hinged_btn_image,borderwidth=1,height=50,width=50,command=lambda : add_fixed_support("hinged"))
    hinge_support.place(x=0,y=210)
    
    command_label5=Label(root,text='Add Hinged support',font=('Helvetica',20),bg='#15ebe7')
    hinge_support.bind('<Enter>',lambda x:

                                    command_label5.place(x=50,y=210)
                                    
                                        )
    hinge_support.bind('<Leave>',lambda x:
                                    command_label5.place_forget()
                                        )

    ##################
    
    
    

############################



def line_btn_for_binding_btn_1():
    my_canvas.bind('<Button-1>',line_draw)
    my_canvas.unbind('<Motion>')
    try:
        my_canvas.delete(show_window)
    except:
        pass
    
    

root=Tk()
global c,counter
c=0
counter=0
w=10000
h=5000
gr=int(input("grid spaces = "))
off=gr*10

my_canvas=Canvas(root,width=w,height=h,bg='white',cursor="crosshair",background='#DAEAC6')
my_canvas.place(x=50,y=50)










#create scroll bar
my_scrollbar=ttk.Scrollbar(root,orient=HORIZONTAL,command=my_canvas.xview)


my_scrollbar1=ttk.Scrollbar(root,orient=VERTICAL,command=my_canvas.yview)





#config canvas
my_canvas.configure(xscrollcommand=my_scrollbar.set)
my_canvas.configure(yscrollcommand=my_scrollbar1.set)






import time
global o_x,o_y
o_x=0
o_y=0
def fn2(event):
    print(event.keysym)
    global o_x,o_y
    #time.sleep(1)
    if event.keysym=='Left':
        my_canvas.xview_scroll(-1,'units')
        o_x=o_x-100*10
    if event.keysym=='Right':
        my_canvas.xview_scroll(1,'units')
        o_x=o_x+100*10
    if event.keysym=='Up':
        my_canvas.yview_scroll(-1,'units')
        o_y=o_y-50*10
    if event.keysym=='Down':
        my_canvas.yview_scroll(1,'units')
        o_y=o_y+50*10
    #print(f'o_x = {o_x}   o_y ={o_y}')
root.bind('<Key>',fn2)

#scrolling ends here  ------------------------new root is "second_frame"-----------------------------




try:
    wb=load_workbook("data.xlsx")
    
    
except:
    
    wb=Workbook()
  
sheet=wb["Sheet"]
wb.create_sheet('inertia')
wb.create_sheet('omega')
wb.create_sheet('s_x')
wb.create_sheet('s_y')
wb.create_sheet('hz_loads_leftward')
wb.create_sheet('hz_loads_rightward')

inertia_sheet=wb['inertia']
omega_sheet=wb['omega']
sup_x_sheet=wb['s_x']
sup_y_sheet=wb['s_y']
hz_loads_leftward=wb['hz_loads_leftward']
hz_loads_rightward=wb['hz_loads_rightward']





for i in range(0,int(w/off)+1):
    if i==0:
        my_canvas.create_text(off*i+5,8,text=f'{int(off*i/10)}',fill='red')
    else:
        my_canvas.create_text(off*i,8,text=f'{int(off*i/10)}',fill='red')
        
        

    my_canvas.create_line(i*off,12,i*off,h,width=1,fill='#6E7961')
    

for i in range(0,int(h/off)+1):
    if i==0:
        pass
    else:
        my_canvas.create_text(15,off*i,text=f'{int(off*i/10)}',fill='red')

    my_canvas.create_line(20,i*off,w,i*off,width=1,fill='#6E7961')


def analyze_select_member1(event):
    s_or_b='s'
    analyze(event.x+o_x,event.y+o_y,s_or_b)


def analyze_select_member(event):

    s_or_b='b'
    analyze(event.x+o_x,event.y+o_y,s_or_b)





def analyze_binding():
    my_canvas.bind("<Button-1>",analyze_select_member)
    my_canvas.bind("<Button-3>",analyze_select_member1)
    




fixed_btn_image=PhotoImage(file='beams/fix_3.png')
hinged_btn_image=PhotoImage(file='beams/hin_1.png') 
point_btn_image=PhotoImage(file='beams/point_btn.png')
save_btn_image=PhotoImage(file='beams/build.png')



line_image=PhotoImage(file='beams/line.png')
analyze_image=PhotoImage(file='beams/run.png')
color_image=PhotoImage(file='beams/color.png')
###############
line_btn=Button(root,text='Line',command=line_btn_for_binding_btn_1,image=line_image,borderwidth=1,height=50,width=50)
line_btn.place(x=0,y=60)
#################
command_label=Label(root,text='Line',font=('Helvetica',20),bg='#15ebe7')
line_btn.bind('<Enter>',lambda x:

                                command_label.place(x=50,y=50)
                                
                                    )
line_btn.bind('<Leave>',lambda x:
                                command_label.place_forget()
                                    )

##################
analyze_btn=Button(root,text='Analyze',command=analyze_binding,image=analyze_image,borderwidth=1,height=50,width=80)
analyze_btn.place(x=50,y=0)
#################
command_label1=Label(root,text='Analyze',font=('Helvetica',20),bg='#15ebe7')
analyze_btn.bind('<Enter>',lambda x:

                                command_label1.place(x=50,y=80)
                                
                                    )
analyze_btn.bind('<Leave>',lambda x:
                                command_label1.place_forget()
                                    )

##################

global color_p
def color_picker():
    
    color_p=colorchooser.askcolor()[1]
    my_canvas.config(bg=color_p)
color_btn=Button(root,text='',command=color_picker,image=color_image,borderwidth=1,height=50,width=80)
color_btn.place(x=230,y=0)

##################

command_label2=Label(root,text='Choose background color',font=('Helvetica',20),bg='#15ebe7')
color_btn.bind('<Enter>',lambda x:

                                command_label2.place(x=230,y=80)
                                
                                    )
color_btn.bind('<Leave>',lambda x:
                                command_label2.place_forget()
                                    )

##################



root.mainloop()

